# 棒グラフ #
from openpyxl import load_workbook
from openpyxl.chart import BarChart, reference
wb = load_workbook(filename='output.xlsx')
sheet_name = wb.sheetnames[0]

x = reference(ws1, min_col=1, min_row=2, max_col=3, min_low=13)
data = reference(ws1, min_col=1, min_row=2, max_col=3, min_low=13)
chart = BarChart()
chart.add_data(data)
chart.set_categories(x)
ws1.add_chart(chart, 'E3')
wb.save('output.xlsx')