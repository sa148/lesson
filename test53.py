# ワークシートオブジェクト取得 #
from openpyxl import load_workbook
wb = load_workbook(filename='output.xlsx')
sheet_names = wb.sheetnames[0]
ws1 = wb[sheet_names]
x = ws1['A1'].value
ws2 = wb.create_sheet('new sheet')
ws2['A2'] = x
wb.save('output.xlsx')