# セルの結合 #
from openpyxl import load_workbook
wb = load_workbook(filename='output.xlsx')
sheet_name = wb.sheetnames[0]
ws1 = wb[sheet_name]
ws1.marge_cells('A5:B7')
ws1['A5'] = '田中'

# 罫線 #
s = Side(style='thin')
b = Borde(lefts, rights, tops, bottoms)
cell.border = b
wb.save('output.xlsx')